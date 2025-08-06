import streamlit as st
import pandas as pd
import os
import re
from collections import defaultdict
from io import BytesIO
import plotly.express as px
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import smtplib
from email.message import EmailMessage

# Configuratie
BEHEERDER_WACHTWOORD = "Beheer123"

GOUD = "FFD700"
ZILVER = "C0C0C0"
BRONS = "CD7F32"
ROOD = "FF9999"

PUNTEN_PER_PLAATS = [10,9,8,7,6,5,4,3,2,1]
MAX_DAGPRIJZEN = 2

RENNERS_FILE = "renners.xlsx"
DEELNEMERS_FILE = "deelnemers.xlsx"
ETAPPES_FILE = "etappes.xlsx"

def check_bestanden():
    ontbreekt = []
    for f in [RENNERS_FILE, DEELNEMERS_FILE, ETAPPES_FILE]:
        if not os.path.exists(f):
            ontbreekt.append(f)
    return ontbreekt

def kleur_cel(cell, kleur):
    cell.fill = PatternFill(start_color=kleur, end_color=kleur, fill_type="solid")

def bereken_klassement():
    ontbreekt = check_bestanden()
    if ontbreekt:
        return None, None, None, None, f"⚠️ Ontbrekende bestanden: {', '.join(ontbreekt)}"

    df_deelnemers = pd.read_excel(DEELNEMERS_FILE)
    df_etappes = pd.read_excel(ETAPPES_FILE)
    df_renners = pd.read_excel(RENNERS_FILE)
    
    deelnemers = {}
    for _, rij in df_deelnemers.iterrows():
        naam = rij["Naam"]
        deelnemers[naam] = {
            "teamnaam": rij.get("Teamnaam", ""),
            "mail": rij.get("Mailadres", ""),
            "adres": rij.get("Thuisadres", ""),
            "telefoon": rij.get("Telefoonnummer", ""),
            "bank": rij.get("Bankrekeningnummer", ""),
            "renners": [rij[f"R{i}"] for i in range(1,11)],
            "reserves": [rij[f"Res{i}"] for i in range(1,6)],
            "actief": [rij[f"R{i}"] for i in range(1,11)],
            "scores": [],
        }
    
    uitvallers_per_dag = defaultdict(set)
    dagprijzen = defaultdict(int)
    dagwinnaars_lijst = []
    punten_per_renner = defaultdict(int)

    for _, row in df_etappes.iterrows():
        dag = int(row["Dag"])
        top10 = [row[f"Top{i}"] for i in range(1,11)]
        
        dns_col = f"DNS Dag {dag}"
        dnf_col = f"DNF Dag {dag}"

        if dag > 1:
            vorige_uitvallers = uitvallers_per_dag[dag-1]
            for deelnemer, data in deelnemers.items():
                nieuwe_actief = []
                reserves_copy = data["reserves"][:]
                for renner in data["actief"]:
                    renners_row = df_renners[df_renners["Rugnummer"] == renner]
                    is_dns = False
                    is_dnf = False
                    if not renners_row.empty:
                        if dns_col in df_renners.columns and pd.notna(renners_row.iloc[0].get(dns_col)):
                            is_dns = str(renners_row.iloc[0][dns_col]).strip().upper() == "X"
                        if dnf_col in df_renners.columns and pd.notna(renners_row.iloc[0].get(dnf_col)):
                            is_dnf = str(renners_row.iloc[0][dnf_col]).strip().upper() == "X"
                    if renner in vorige_uitvallers or is_dns or is_dnf:
                        if reserves_copy:
                            nieuwe_actief.append(reserves_copy.pop(0))
                        else:
                            pass
                    else:
                        nieuwe_actief.append(renner)
                data["actief"] = nieuwe_actief

        dagscores = {}
        for deelnemer, data in deelnemers.items():
            score = sum(PUNTEN_PER_PLAATS[i] for i, renner in enumerate(top10) if renner in data["actief"])
            data["scores"].append(score)
            dagscores[deelnemer] = score
            for i, renner in enumerate(top10):
                if renner in data["actief"]:
                    punten_per_renner[renner] += PUNTEN_PER_PLAATS[i]

        max_score = max(dagscores.values()) if dagscores else 0
        kandidaten = [n for n, s in dagscores.items() if s == max_score]

        if len(kandidaten) > 1:
            totaalscores_tot_vorig = {n: sum(deelnemers[n]["scores"][:-1]) for n in kandidaten}
            laagste_score = min(totaalscores_tot_vorig.values())
            winnaars = [n for n, ts in totaalscores_tot_vorig.items() if ts == laagste_score]
            if len(winnaars) > 1:
                for w in winnaars:
                    if dagprijzen[w] < MAX_DAGPRIJZEN:
                        dagprijzen[w] += 1
            else:
                w = winnaars[0]
                if dagprijzen[w] < MAX_DAGPRIJZEN:
                    dagprijzen[w] += 1
        else:
            w = kandidaten[0]
            if dagprijzen[w] < MAX_DAGPRIJZEN:
                dagprijzen[w] += 1

        if len(kandidaten) == 1:
            dagwinnaars_lijst.append(f"Dag {dag}: {kandidaten[0]} ({max_score} punten)")
        else:
            dagwinnaars_lijst.append(f"Dag {dag}: {' & '.join(kandidaten)} ({max_score} punten)")

        uitvallers_per_dag[dag] = set()
        for _, data in deelnemers.items():
            for renner in data["actief"]:
                renners_row = df_renners[df_renners["Rugnummer"] == renner]
                is_dns = False
                is_dnf = False
                if not renners_row.empty:
                    if dns_col in df_renners.columns and pd.notna(renners_row.iloc[0].get(dns_col)):
                        is_dns = str(renners_row.iloc[0][dns_col]).strip().upper() == "X"
                    if dnf_col in df_renners.columns and pd.notna(renners_row.iloc[0].get(dnf_col)):
                        is_dnf = str(renners_row.iloc[0][dnf_col]).strip().upper() == "X"
                if is_dns or is_dnf:
                    uitvallers_per_dag[dag].add(renner)

    totaal_scores = {n: sum(d["scores"]) for n, d in deelnemers.items()}
    max_totaal = max(totaal_scores.values()) if totaal_scores else 0

    data_out = []
    for naam, d in deelnemers.items():
        row = {
            "Naam": naam, "Teamnaam": d["teamnaam"],
            "Mailadres": d["mail"], "Thuisadres": d["adres"],
            "Telefoonnummer": d["telefoon"], "Bankrekeningnummer": d["bank"]
        }
        for i, score in enumerate(d["scores"], start=1):
            row[f"Dag {i}"] = score
        row["Totaal"] = totaal_scores[naam]
        row["Verschil"] = max_totaal - totaal_scores[naam]
        row["Dagprijzen"] = dagprijzen[naam]
        data_out.append(row)

    kolommen = [
        "Naam", "Teamnaam", "Mailadres", "Thuisadres", "Telefoonnummer", "Bankrekeningnummer"
    ] + [f"Dag {i}" for i in range(1, len(df_etappes)+1)] + ["Totaal", "Verschil", "Dagprijzen"]
    df_resultaat = pd.DataFrame(data_out).reindex(columns=kolommen)
    df_resultaat = df_resultaat.sort_values(by="Totaal", ascending=False).reset_index(drop=True)

    bestand = "klassement.xlsx"
    df_resultaat.to_excel(bestand, index=False)
    wb = load_workbook(bestand)
    ws = wb.active

    for dag_i in range(7, 7 + len(df_etappes)):
        scores = [(row, ws.cell(row=row, column=dag_i).value) for row in range(2, ws.max_row + 1)]
        max_score_dag = max(v for _, v in scores if isinstance(v, (int, float)))
        for row, val in scores:
            if val == max_score_dag:
                kleur_cel(ws.cell(row=row, column=dag_i), GOUD)

    wb.save(bestand)

    return df_resultaat, dagwinnaars_lijst, punten_per_renner, bestand, None


def verstuur_email(smtp_server, smtp_port, smtp_email, smtp_wachtwoord, 
                   ontvangers, onderwerp, bericht):
    try:
        msg = EmailMessage()
        msg["From"] = smtp_email
        msg["To"] = ", ".join(ontvangers)
        msg["Subject"] = onderwerp
        msg.set_content(bericht)

        with smtplib.SMTP_SSL(smtp_server, smtp_port) as smtp:
            smtp.login(smtp_email, smtp_wachtwoord)
            smtp.send_message(msg)

        return True, "Mail succesvol verstuurd!"
    except Exception as e:
        return False, f"Fout bij versturen mail: {e}"


st.set_page_config(page_title="Tourspel", layout="wide")
st.title("🚴 Tour de France Tourspel")

ontbreekt = check_bestanden()
if ontbreekt:
    st.warning(f"⚠️ Ontbrekende bestanden: {', '.join(ontbreekt)}. Voeg ze toe of update via Inschrijven.")

if "beheerder_ingelogd" not in st.session_state:
    st.session_state["beheerder_ingelogd"] = False

def beheerder_login():
    st.sidebar.header("🔐 Beheerder Login")
    if not st.session_state["beheerder_ingelogd"]:
        wachtwoord = st.sidebar.text_input("Wachtwoord", type="password")
        if st.sidebar.button("Inloggen"):
            if wachtwoord == BEHEERDER_WACHTWOORD:
                st.session_state["beheerder_ingelogd"] = True
                st.sidebar.success("Inloggen geslaagd!")
            else:
                st.sidebar.error("Onjuist wachtwoord.")
    else:
        if st.sidebar.button("Uitloggen"):
            st.session_state["beheerder_ingelogd"] = False
            st.sidebar.info("Uitgelogd.")

beheerder_login()
if st.session_state["beheerder_ingelogd"]:
    st.subheader("⬆️ Upload nieuwe bestanden")

    def upload_en_opslaan(bestandsnaam, label):
        upload = st.file_uploader(label, type=["xlsx"], key=bestandsnaam)
        if upload is not None:
            try:
                df = pd.read_excel(upload)
                df.to_excel(bestandsnaam, index=False)
                st.success(f"{label} succesvol geüpload en opgeslagen als {bestandsnaam}")
            except Exception as e:
                st.error(f"Fout bij uploaden van {label}: {e}")

    upload_en_opslaan(RENNERS_FILE, "Renners bestand (renners.xlsx)")
    upload_en_opslaan(DEELNEMERS_FILE, "Deelnemers bestand (deelnemers.xlsx)")
    upload_en_opslaan(ETAPPES_FILE, "Etappes bestand (etappes.xlsx)")

# Inschrijven knop en formulier
if st.button("📝 Inschrijven"):
    if not os.path.exists(RENNERS_FILE):
        st.warning("⚠️ Rennersbestand ontbreekt.")
    else:
        df_renners = pd.read_excel(RENNERS_FILE)
        opties = [f"{row['Rugnummer']} – {row['Naam']}" for _, row in df_renners.iterrows()]

        with st.form("inschrijfformulier"):
            naam = st.text_input("Naam deelnemer", max_chars=50)
            mail = st.text_input("E-mail adres", max_chars=100)
            adres = st.text_area("Thuisadres", max_chars=250)
            telefoon = st.text_input("Telefoonnummer", max_chars=20)
            bank = st.text_input("Bankrekeningnummer", max_chars=34)
            teamnaam = st.text_input("Teamnaam", max_chars=50)

            hoofdrenners = []
            reserves = []
            st.write("**Hoofdrenners (10):**")
            for i in range(1,11):
                hoofdrenners.append(st.selectbox(f"Hoofdrenner {i}", opties, key=f"h{i}"))
            st.write("**Reserverenners (5):**")
            for i in range(1,6):
                reserves.append(st.selectbox(f"Reserverenner {i}", opties, key=f"r{i}"))

            submit = st.form_submit_button("💾 Opslaan ploeg")

            if submit:
                gekozen = hoofdrenners + reserves
                nummers = [int(opt.split(" – ")[0]) for opt in gekozen]
                if len(set(nummers)) != 15:
                    st.error("❌ Kies 15 unieke renners (hoofd + reserve).")
                elif not naam.strip() or not mail.strip() or not adres.strip() or not telefoon.strip() or not bank.strip() or not teamnaam.strip():
                    st.error("❌ Vul alle velden in.")
                else:
                    kolommen = ["Naam", "Teamnaam", "Mailadres", "Thuisadres", "Telefoonnummer", "Bankrekeningnummer"] + [f"R{i}" for i in range(1,11)] + [f"Res{i}" for i in range(1,6)]
                    if os.path.exists(DEELNEMERS_FILE):
                        df_deelnemers = pd.read_excel(DEELNEMERS_FILE)
                        if naam in df_deelnemers["Naam"].values:
                            df_deelnemers.loc[df_deelnemers["Naam"] == naam, kolommen] = [naam, teamnaam, mail, adres, telefoon, bank] + nummers
                            st.success(f"✅ Ploeg van {naam} bijgewerkt.")
                        else:
                            nieuwe_rij = pd.DataFrame([[naam, teamnaam, mail, adres, telefoon, bank] + nummers], columns=kolommen)
                            df_deelnemers = pd.concat([df_deelnemers, nieuwe_rij], ignore_index=True)
                            st.success(f"✅ Ploeg van {naam} toegevoegd.")
                    else:
                        df_deelnemers = pd.DataFrame([[naam, teamnaam, mail, adres, telefoon, bank] + nummers], columns=kolommen)
                        st.success(f"✅ Ploeg van {naam} toegevoegd.")

                    df_deelnemers.to_excel(DEELNEMERS_FILE, index=False)

# Tabs
tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📊 Klassement",
    "🏁 Etappe-uitslagen",
    "🏅 Dagwinnaars",
    "🚴 Punten per renner",
    "📈 Statistieken"
])

with tab2:
    st.header("Klassement")
    if st.button("📊 Bereken klassement"):
        resultaat, dagwinnaars, punten_per_renner, bestand, fout = bereken_klassement()
        if fout:
            st.error(fout)
        else:
            st.dataframe(resultaat.drop(columns=["Mailadres", "Thuisadres", "Telefoonnummer", "Bankrekeningnummer"]), use_container_width=True)
            with open(bestand, "rb") as f:
                st.download_button("📥 Download klassement.xlsx", f, file_name="klassement.xlsx")
            st.session_state["klassement"] = resultaat
            st.session_state["dagwinnaars"] = dagwinnaars
            st.session_state["punten_per_renner"] = punten_per_renner

with tab3:
    st.header("Etappe-uitslagen")
    if not os.path.exists(ETAPPES_FILE) or not os.path.exists(RENNERS_FILE):
        st.warning("⚠️ Renners- of etappesbestand ontbreekt.")
    else:
        df_etappes = pd.read_excel(ETAPPES_FILE)
        df_renners = pd.read_excel(RENNERS_FILE)
        etappe_opties = df_etappes["Dag"].astype(str).tolist()
        etappe_keuze = st.selectbox("Selecteer etappe", options=etappe_opties)
        uitval_optie = st.checkbox("Toon uitvallers bij etappe", value=True)

        rij = df_etappes[df_etappes["Dag"] == int(etappe_keuze)].iloc[0]
        top10 = [rij[f"Top{i}"] for i in range(1,11)]
        uitvallers_str = str(rij["Uitvallers"]) if pd.notna(rij["Uitvallers"]) else ""
        uitvallers = list(map(int, re.split(r"\s*,\s*", uitvallers_str))) if uitvallers_str else []

        def renner_info(rugnr):
            rijr = df_renners[df_renners["Rugnummer"] == rugnr]
            if rijr.empty:
                return ("Onbekend", "Onbekend")
            return (rijr["Naam"].values[0], rijr["Team"].values[0])

        st.subheader(f"Etappe {etappe_keuze} - Top 10")
        for i, rugnr in enumerate(top10, 1):
            naam, team = renner_info(rugnr)
            st.write(f"{i}. {naam} ({rugnr}) - {team}")

        if uitval_optie:
            st.subheader("Uitvallers")
            if uitvallers:
                for rugnr in uitvallers:
                    naam, team = renner_info(rugnr)
                    st.write(f"{naam} ({rugnr}) - {team}")
            else:
                st.write("Geen uitvallers.")

with tab4:
    st.header("Dagwinnaars")
    if "dagwinnaars" in st.session_state:
        for dw in st.session_state["dagwinnaars"]:
            st.write(dw)
    else:
        st.info("ℹ️ Bereken eerst het klassement om dagwinnaars te zien.")

with tab5:
    st.header("Punten per renner")
    if not os.path.exists(RENNERS_FILE):
        st.warning("⚠️ Rennersbestand ontbreekt.")
    else:
        df_renners = pd.read_excel(RENNERS_FILE)
        punten_per_renner = st.session_state.get("punten_per_renner")
        if not punten_per_renner:
            st.info("ℹ️ Bereken eerst het klassement om punten te tonen.")
        else:
            data = []
            for rugnr, punten in punten_per_renner.items():
                rij = df_renners[df_renners["Rugnummer"] == rugnr]
                naam = rij["Naam"].values[0] if not rij.empty else "Onbekend"
                team = rij["Team"].values[0] if not rij.empty else "Onbekend"
                data.append({"Rugnummer": rugnr, "Naam": naam, "Team": team, "Punten": punten})
            df_punten = pd.DataFrame(data)
            zoekterm = st.text_input("Zoek op naam of rugnummer")
            if zoekterm.strip():
                zoekterm_lc = zoekterm.lower()
                df_punten = df_punten[
                    df_punten["Naam"].str.lower().str.contains(zoekterm_lc) |
                    df_punten["Rugnummer"].astype(str).str.contains(zoekterm_lc)
                ]
            teams = ["Alle teams"] + sorted(df_punten["Team"].unique())
            team_filter = st.selectbox("Filter op team", teams)
            if team_filter != "Alle teams":
                df_punten = df_punten[df_punten["Team"] == team_filter]
            df_punten = df_punten.sort_values(by="Punten", ascending=False).reset_index(drop=True)
            st.dataframe(df_punten, use_container_width=True)

with tab6:
    st.header("Statistieken")
    if not os.path.exists(RENNERS_FILE) or not os.path.exists(DEELNEMERS_FILE):
        st.warning("⚠️ Renners- of deelnemersbestand ontbreekt.")
    else:
        df_klassement = st.session_state.get("klassement")
        if df_klassement is None:
            st.info("ℹ️ Bereken eerst het klassement om statistieken te zien.")
        else:
            df_long = df_klassement.melt(
                id_vars=["Naam", "Teamnaam", "Totaal", "Verschil", "Dagprijzen"],
                value_vars=[c for c in df_klassement.columns if c.startswith("Dag")],
                var_name="Etappe",
                value_name="Punten"
            )
            fig = px.line(df_long, x="Etappe", y="Punten", color="Naam", markers=True, title="Puntenontwikkeling per deelnemer")
            st.plotly_chart(fig, use_container_width=True)
            fig2 = px.bar(df_klassement, x="Naam", y="Dagprijzen", title="Aantal dagprijzen per deelnemer",
                          color="Dagprijzen", color_continuous_scale="YlOrBr")
            st.plotly_chart(fig2, use_container_width=True)
            teampunten = df_klassement.groupby("Teamnaam")["Totaal"].sum().reset_index()
            teampunten = teampunten.sort_values(by="Totaal", ascending=False)
            fig3 = px.bar(teampunten, x="Teamnaam", y="Totaal", title="Totaal punten per team",
                          color="Totaal", color_continuous_scale="Blues")
            st.plotly_chart(fig3, use_container_width=True)

if st.session_state["beheerder_ingelogd"]:
    st.sidebar.header("🛠️ Beheerdersdashboard")
    st.sidebar.info("Je bent ingelogd als beheerder. Je kunt alle deelnemersgegevens zien, inclusief privégegevens.")
    if os.path.exists(DEELNEMERS_FILE):
        df_deelnemers = pd.read_excel(DEELNEMERS_FILE)
        st.sidebar.write(f"Aantal deelnemers: {len(df_deelnemers)}")
        zoekterm = st.sidebar.text_input("Zoek deelnemer op naam of mailadres")
        if zoekterm.strip():
            zoek_lc = zoekterm.lower()
            df_show = df_deelnemers[
                df_deelnemers["Naam"].str.lower().str.contains(zoek_lc) |
                df_deelnemers["Mailadres"].str.lower().str.contains(zoek_lc)
            ]
        else:
            df_show = df_deelnemers
        st.sidebar.dataframe(df_show, use_container_width=True)
    else:
        st.sidebar.warning("Geen deelnemersbestand gevonden.")

    st.subheader("📧 Mail naar deelnemers")
    smtp_server = st.text_input("SMTP-server (bijv. smtp.gmail.com)")
    smtp_port = st.number_input("SMTP-poort", value=465)
    smtp_email = st.text_input("Jouw e-mailadres (afzender)")
    smtp_wachtwoord = st.text_input("Wachtwoord", type="password")

    if os.path.exists(DEELNEMERS_FILE):
        opties = df_deelnemers["Mailadres"].dropna().tolist()
        ontvangers = st.multiselect("Selecteer ontvangers", opties, default=opties)

        onderwerp = st.text_input("Onderwerp")
        bericht = st.text_area("Bericht")

        if st.button("Verstuur mail"):
            if not (smtp_server and smtp_port and smtp_email and smtp_wachtwoord and onderwerp and bericht):
                st.error("Vul alle velden in!")
            elif not ontvangers:
                st.error("Selecteer minimaal één ontvanger!")
            else:
                success, message = verstuur_email(smtp_server, smtp_port, smtp_email, smtp_wachtwoord, ontvangers, onderwerp, bericht)
                if success:
                    st.success(message)
                else:
                    st.error(message)
    else:
        st.info("Geen deelnemersbestand gevonden.")
